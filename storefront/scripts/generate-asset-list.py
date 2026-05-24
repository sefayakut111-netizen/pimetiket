"""Pim Etiket — tasarımcı için PNG/SVG asset listesi (xlsx)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BRAND_RED = "FF6B5B"
NAVY = "1F2937"
TINT = "FFF1EE"

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color=NAVY)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="4B5563")
BODY_FONT = Font(name="Arial", size=10)
BRAND_FILL = PatternFill("solid", start_color=TINT)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def setup_sheet(ws, title, subtitle, headers, col_widths):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A3"] = (
        "Brand: Coral #FF6B5B | Navy #1F2937 | Cream #F5EBD9 | Tint #FFF1EE  ||  "
        "Mascot: Pim (stilize karga)  ||  Stil: Flat + subtle neumorphism, soft shadows  ||  "
        "Retina: PNG 2x boyutta uret, SVG vector kalir"
    )
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color=BRAND_RED)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 28
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A6"


def add_row(ws, row_num, values, fill=None):
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[row_num].height = 32


# ============================================================
# SHEET 1: URUN RESIMLERI (22 adet)
# ============================================================
ws1 = wb.active
ws1.title = "Urun Resimleri (22)"
setup_sheet(
    ws1,
    "Urun Resimleri — 22 adet",
    "Etiket grid (11) + Sticker grid (11) kart gorselleri",
    [
        "No",
        "Kategori",
        "Turkce Adi",
        "Ingilizce Adi",
        "Dosya Adi",
        "Boyut (px)",
        "Format",
        "Slot Path",
        "Aciklama",
    ],
    [5, 14, 22, 22, 24, 12, 8, 26, 50],
)

urunler = [
    # ETIKET RULO (6)
    [1, "Etiket / Rulo", "Ozel kesim rulo", "Die-cut roll", "rulo-diecut.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + die-cut blob siluet, dashed kesim cizgisi"],
    [2, "Etiket / Rulo", "Seffaf rulo", "Clear roll", "rulo-clear.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + cam gorunumlu transparan etiket"],
    [3, "Etiket / Rulo", "Yuvarlak rulo", "Circle roll", "rulo-circle.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + 2 daire stack (back soluk, on canli)"],
    [4, "Etiket / Rulo", "Kare rulo", "Square roll", "rulo-square.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + 2 kare stack (slight rotation)"],
    [5, "Etiket / Rulo", "Dikdortgen rulo", "Rectangle roll", "rulo-rectangle.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + 2 yatay dikdortgen stack"],
    [6, "Etiket / Rulo", "Oval rulo", "Oval roll", "rulo-oval.png", "220x130", "PNG", "public/etiket-cards/", "Yatik silindir + 2 oval/elips stack"],
    # ETIKET TABAKA (5)
    [7, "Etiket / Tabaka", "Yuvarlak tabaka", "Circle sheet", "tabaka-circle.png", "220x130", "PNG", "public/etiket-cards/", "Dikey A4 kagit + 3x4 = 12 daire grid"],
    [8, "Etiket / Tabaka", "Ozel kesim tabaka", "Die-cut sheet", "tabaka-diecut.png", "220x130", "PNG", "public/etiket-cards/", "Dikey A4 kagit + 6 Pim mascot silueti (2x3 grid)"],
    [9, "Etiket / Tabaka", "Oval tabaka", "Oval sheet", "tabaka-oval.png", "220x130", "PNG", "public/etiket-cards/", "Dikey A4 kagit + 2x4 = 8 oval grid"],
    [10, "Etiket / Tabaka", "Dikdortgen tabaka", "Rectangle sheet", "tabaka-rectangle.png", "220x130", "PNG", "public/etiket-cards/", "Dikey A4 kagit + 2x4 = 8 dikdortgen grid"],
    [11, "Etiket / Tabaka", "Kare tabaka", "Square sheet", "tabaka-square.png", "220x130", "PNG", "public/etiket-cards/", "Dikey A4 kagit + 3x3 = 9 kare grid"],
    # STICKER (11)
    [12, "Sticker", "Ozel kesim sticker", "Die cut sticker", "diecut.png", "200x130", "PNG", "public/sticker-cards/", "Organik blob die-cut siluet, dashed kesim cizgisi"],
    [13, "Sticker", "Yuvarlak sticker", "Circle sticker", "circle.png", "200x130", "PNG", "public/sticker-cards/", "2 stacked daire (back soluk, on canli)"],
    [14, "Sticker", "Dikdortgen sticker", "Rectangle sticker", "rectangle.png", "200x130", "PNG", "public/sticker-cards/", "2 stacked dikdortgen, sharp veya rounded corner"],
    [15, "Sticker", "Kare sticker", "Square sticker", "square.png", "200x130", "PNG", "public/sticker-cards/", "2 stacked kare, sharp veya rounded corner"],
    [16, "Sticker", "Oval sticker", "Oval sticker", "oval.png", "200x130", "PNG", "public/sticker-cards/", "2 stacked oval/elips"],
    [17, "Sticker", "Bumper sticker", "Bumper sticker", "bumper.png", "200x130", "PNG", "public/sticker-cards/", "Uzun-yatay pill formu (otomobil tampon stickeri)"],
    [18, "Sticker", "Yari kesim sticker", "Kiss cut sticker", "kisscut.png", "200x130", "PNG", "public/sticker-cards/", "Beyaz backing paper + sticker siluet, kesim sadece sticker ustunde"],
    [19, "Sticker", "Seffaf sticker", "Clear sticker", "clear.png", "200x130", "PNG", "public/sticker-cards/", "Cam berraklik transparan siluet, hafif mavi tint"],
    [20, "Sticker", "Holografik sticker", "Holographic sticker", "holo.png", "200x130", "PNG", "public/sticker-cards/", "Gokkusagi prizmatik yansima (pembe/mor/cyan/yesil/sari)"],
    [21, "Sticker", "Simli sticker", "Glitter sticker", "simli.png", "200x130", "PNG", "public/sticker-cards/", "Pariltili metalik dokulu (beyaz dot pattern)"],
    [22, "Sticker", "Sticker sayfasi", "Sticker sheet", "sheets.png", "200x130", "PNG", "public/sticker-cards/", "A4 kagit + karma sekiller grid (mixed circles/squares/ovals)"],
]
for i, row in enumerate(urunler):
    add_row(ws1, 6 + i, row, fill=BRAND_FILL if i % 2 == 0 else None)


# ============================================================
# SHEET 2: SARIM YONU (8)
# ============================================================
ws2 = wb.create_sheet("Sarim Yonu (8)")
setup_sheet(
    ws2,
    "Sarim Yonu Ikonlari — 8 adet",
    "Etiket rulo icin sarim yonu ikonlari (4 disa + 4 ice)",
    ["No", "Tur", "Yon No", "Aciklama", "Dosya Adi", "Boyut (px)", "Format", "Notlar"],
    [5, 14, 8, 32, 26, 12, 8, 36],
)

sarim = [
    [1, "Disa sarim", 1, "En yaygin kullanilan — etiket disa bakar, ustten saga", "winding-out-1.svg", "64x64", "SVG", "PopulerBadge ile isaretli (varsayilan secim)"],
    [2, "Disa sarim", 2, "Disa bakar, ustten sola", "winding-out-2.svg", "64x64", "SVG", "90 derece dondurulmus versiyon"],
    [3, "Disa sarim", 3, "Disa bakar, alttan saga", "winding-out-3.svg", "64x64", "SVG", "180 derece dondurulmus versiyon"],
    [4, "Disa sarim", 4, "Disa bakar, alttan sola", "winding-out-4.svg", "64x64", "SVG", "270 derece dondurulmus versiyon"],
    [5, "Ice sarim", 1, "Ice bakar — etiket yuzu ice gelir, ustten saga", "winding-in-1.svg", "64x64", "SVG", "Ice sarim varyanti (mirror)"],
    [6, "Ice sarim", 2, "Ice bakar, ustten sola", "winding-in-2.svg", "64x64", "SVG", "Mirror + 90 derece rotation"],
    [7, "Ice sarim", 3, "Ice bakar, alttan saga", "winding-in-3.svg", "64x64", "SVG", "Mirror + 180 derece rotation"],
    [8, "Ice sarim", 4, "Ice bakar, alttan sola", "winding-in-4.svg", "64x64", "SVG", "Mirror + 270 derece rotation"],
]
for i, row in enumerate(sarim):
    add_row(ws2, 6 + i, row, fill=BRAND_FILL if i % 2 == 0 else None)


# ============================================================
# SHEET 3: ETIKET MALZEME (6)
# ============================================================
ws3 = wb.create_sheet("Etiket Malzeme (6)")
setup_sheet(
    ws3,
    "Etiket Malzeme Ikonlari — 6 adet",
    "Rulo + tabaka icin malzeme yuzey ornekleri (Step 1: Malzeme)",
    ["No", "Malzeme ID", "Turkce Adi", "Ingilizce Adi", "Dosya Adi", "Boyut (px)", "Format", "Yuzey Ozelligi", "Aciklama"],
    [5, 14, 22, 22, 22, 12, 8, 24, 42],
)

etiket_malzeme = [
    [1, "kuse", "Kuse Etiket", "Coated paper", "kuse.svg", "220x110", "SVG", "Puruzsuz mat kagit", "Standart baski kagidi, en yaygin secim"],
    [2, "kraft", "Kraft Etiket", "Kraft paper", "kraft.svg", "220x110", "SVG", "Dogal dokulu eko kagit", "Koy/organik urun hissi (su an gizli — partner yok)"],
    [3, "beyaz", "Opak PP Etiket", "Opaque PP", "opakpp.svg", "220x110", "SVG", "Yirtilmaz plastik", "Su/yag/nem dayanikli, soguk zincir urunleri icin"],
    [4, "seffaf", "Seffaf Etiket", "Transparent", "transparan.svg", "220x110", "SVG", "Saydam film", "Arka plan gorunur, cam sise urunlerinde"],
    [5, "ultra", "Ultra Clear Etiket", "Ultra Clear", "ultraclear.svg", "220x110", "SVG", "Cam berrakliginda", "Gorunmez film, sadece basilan tasarim gozukur (su an gizli)"],
    [6, "metalik", "Metalize Etiket", "Metallic", "metalik.svg", "220x110", "SVG", "Parlak metalik yuzey", "Premium gold/silver hissi, kozmetik/icecek icin"],
]
for i, row in enumerate(etiket_malzeme):
    add_row(ws3, 6 + i, row, fill=BRAND_FILL if i % 2 == 0 else None)


# ============================================================
# SHEET 4: STICKER MALZEME (4)
# ============================================================
ws4 = wb.create_sheet("Sticker Malzeme (4)")
setup_sheet(
    ws4,
    "Sticker Malzeme Ikonlari — 4 adet",
    "Sticker icin malzeme yuzey ornekleri (Step 1: Malzeme)",
    ["No", "Malzeme ID", "Turkce Adi", "Ingilizce Adi", "Dosya Adi", "Boyut (px)", "Format", "Aciklama"],
    [5, 14, 18, 18, 22, 12, 8, 52],
)

sticker_malzeme = [
    [1, "vinil", "Vinil", "Vinyl", "vinil.svg", "220x110", "SVG", "Standart vinyl sticker (parlak yuzey, dis mekan dayanikli)"],
    [2, "transparan", "Transparan", "Transparent", "transparan.svg", "220x110", "SVG", "Seffaf vinyl, sadece tasarim gorunur (cam/pencere icin)"],
    [3, "holo", "Holografik", "Holographic", "holografik.svg", "220x110", "SVG", "Gokkusagi prizmatik yansima (premium, dikkat cekici)"],
    [4, "simli", "Simli", "Glitter", "simli.svg", "220x110", "SVG", "Pariltili metalik doku (kid-friendly, etkinlik sticker)"],
]
for i, row in enumerate(sticker_malzeme):
    add_row(ws4, 6 + i, row, fill=BRAND_FILL if i % 2 == 0 else None)


# ============================================================
# SHEET 5: KAPLAMA (7)
# ============================================================
ws5 = wb.create_sheet("Kaplama (7)")
setup_sheet(
    ws5,
    "Kaplama Malzemeleri — 7 adet",
    "Etiket kaplama (Step 2) + Sticker yuzey (Step 2) ikonlari",
    ["No", "Tip", "Kaplama ID", "Turkce Adi", "Ingilizce Adi", "Dosya Adi", "Boyut (px)", "Format", "Aciklama"],
    [5, 10, 14, 18, 22, 22, 12, 8, 42],
)

kaplama = [
    [1, "Etiket", "yok", "Kaplamasiz", "Uncoated", "kaplamasiz.svg", "220x110", "SVG", "Dogal kagit dokusu korunur, ekonomik"],
    [2, "Etiket", "mat", "Mat Selefon", "Matte Lamination", "matselefon.svg", "220x110", "SVG", "Yansimasiz mat film, premium gorunum"],
    [3, "Etiket", "parlak", "Parlak Selefon", "Gloss Lamination", "parlakselefon.svg", "220x110", "SVG", "Parlak koruyucu film, canli renkler"],
    [4, "Etiket", "soft", "Soft Touch", "Soft Touch", "softtouch.svg", "220x110", "SVG", "Kadife dokunma hissi, luks (su an gizli — partner yok)"],
    [5, "Sticker", "yok", "Yok", "None", "kaplamasiz.svg", "220x110", "SVG", "Kaplamasiz vinyl (mevcut etiket SVG ile paylasilir)"],
    [6, "Sticker", "mat", "Mat", "Matte", "matselefon.svg", "220x110", "SVG", "Mat selefon (mevcut etiket SVG ile paylasilir)"],
    [7, "Sticker", "parlak", "Parlak", "Gloss", "parlakselefon.svg", "220x110", "SVG", "Parlak selefon (mevcut etiket SVG ile paylasilir)"],
]
for i, row in enumerate(kaplama):
    add_row(ws5, 6 + i, row, fill=BRAND_FILL if i % 2 == 0 else None)


# ============================================================
# Kaydet
# ============================================================
wb.save("docs/PIM-ETIKET-GORSEL-LISTESI.xlsx")
print("OK: docs/PIM-ETIKET-GORSEL-LISTESI.xlsx olusturuldu")
print(f"  - 5 sheet, toplam {22 + 8 + 6 + 4 + 7} = 47 asset")
